import requests
import csv
from datetime import datetime, timezone, timedelta
import config as conf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

class VKStats:
    def __init__(self, vk_api_key: str, group_id: str | int, api_version: str = "5.236"):
        self.vk_api_key = vk_api_key
        self.group_id = group_id
        self.api_version = api_version
        self.base_url = "https://api.vk.com/method/"

    @staticmethod
    def _to_unix_utc(date_str: str, end_of_day: bool = False) -> int:
        """YYYY-MM-DD -> UnixTime (UTC)."""
        dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        if end_of_day:
            dt = (dt + timedelta(days=1)) - timedelta(seconds=1)
        return int(dt.timestamp())

    def get_stats(self, start_date: str, end_date: str):
        """Получаем статистику за период."""
        url = f"{self.base_url}stats.get"
        params = {
            "access_token": self.vk_api_key,
            "v": self.api_version,
            "group_id": self.group_id,
            "timestamp_from": self._to_unix_utc(start_date, end_of_day=False),
            "timestamp_to": self._to_unix_utc(end_date, end_of_day=True),
            "interval": "day",
        }
        resp = requests.get(url, params=params).json()
        if "error" in resp:
            raise Exception(resp["error"]["error_msg"])
        return resp.get("response", [])

    def get_followers(self) -> int:
        """Текущее количество подписчиков группы."""
        url = f"{self.base_url}groups.getMembers"
        params = {
            "access_token": self.vk_api_key,
            "v": self.api_version,
            "group_id": self.group_id,
        }
        resp = requests.get(url, params=params).json()
        if "error" in resp:
            raise Exception(resp["error"]["error_msg"])
        return int(resp["response"]["count"])

    @staticmethod
    def display_stats(rows) -> None:
        """Вывод статистики в консоль."""
        if not rows:
            print("⚠️ Нет статистики за выбранный период.")
            return

        header = (
            f"{'Date':<12}{'Visitors':>10}{'Views':>10}"
            f"{'Likes':>10}{'Subscribers':>14}"
        )
        print(header)
        print("-" * len(header))

        total_visitors = total_views = total_likes = total_subscribers = 0

        for day in rows:
            v = day.get("visitors", {}) or {}
            a = day.get("activity", {}) or {}

            if day.get("day"):
                date_str = day["day"]
            else:
                ts = int(day.get("period_from", 0))
                date_str = datetime.fromtimestamp(ts, timezone.utc).strftime("%Y-%m-%d")

            visitors = int(v.get("visitors", 0))
            views = int(v.get("views", 0))
            likes = int(a.get("likes", 0))
            subscribed = int(v.get("subscribed", 0))
            unsubscribed = int(v.get("unsubscribed", 0))
            subscribers = subscribed - unsubscribed

            total_visitors += visitors
            total_views += views
            total_likes += likes
            total_subscribers += subscribers

            print(f"{date_str:<12}{visitors:>10}{views:>10}{likes:>10}{subscribers:>14}")

        print("-" * len(header))
        print(f"{'TOTAL':<12}{total_visitors:>10}{total_views:>10}{total_likes:>10}{total_subscribers:>14}")

    @staticmethod
    def save_to_csv(rows, filename="stats.csv"):
        """Сохраняем статистику в CSV."""
        with open(filename, mode="w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(["Date", "Visitors", "Views", "Likes", "Subscribers"])

            total_visitors = total_views = total_likes = total_subscribers = 0

            for day in rows:
                v = day.get("visitors", {}) or {}
                a = day.get("activity", {}) or {}

                if day.get("day"):
                    date_str = day["day"]
                else:
                    ts = int(day.get("period_from", 0))
                    date_str = datetime.fromtimestamp(ts, timezone.utc).strftime("%Y-%m-%d")

                visitors = int(v.get("visitors", 0))
                views = int(v.get("views", 0))
                likes = int(a.get("likes", 0))
                subscribed = int(v.get("subscribed", 0))
                unsubscribed = int(v.get("unsubscribed", 0))
                subscribers = subscribed - unsubscribed

                writer.writerow([date_str, visitors, views, likes, subscribers])

                total_visitors += visitors
                total_views += views
                total_likes += likes
                total_subscribers += subscribers

            writer.writerow(["TOTAL", total_visitors, total_views, total_likes, total_subscribers])

    @staticmethod
    def save_to_excel(rows, filename="stats.xlsx"):
        """Сохраняем статистику в Excel с форматированием."""
        wb = Workbook()
        ws = wb.active
        ws.title = "VK Stats"

        headers = ["Date", "Visitors", "Views", "Likes", "Subscribers"]
        ws.append(headers)

        # Заголовки жирные
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        total_visitors = total_views = total_likes = total_subscribers = 0

        for day in rows:
            v = day.get("visitors", {}) or {}
            a = day.get("activity", {}) or {}

            if day.get("day"):
                date_str = day["day"]
            else:
                ts = int(day.get("period_from", 0))
                date_str = datetime.fromtimestamp(ts, timezone.utc).strftime("%Y-%m-%d")

            visitors = int(v.get("visitors", 0))
            views = int(v.get("views", 0))
            likes = int(a.get("likes", 0))
            subscribed = int(v.get("subscribed", 0))
            unsubscribed = int(v.get("unsubscribed", 0))
            subscribers = subscribed - unsubscribed

            ws.append([date_str, visitors, views, likes, subscribers])

            total_visitors += visitors
            total_views += views
            total_likes += likes
            total_subscribers += subscribers

        # Итоговая строка
        total_row = ["TOTAL", total_visitors, total_views, total_likes, total_subscribers]
        ws.append(total_row)

        last_row = ws.max_row
        for col in range(1, len(total_row) + 1):
            ws.cell(row=last_row, column=col).font = Font(bold=True)

        # Автоширина
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        # Подсветка лайков зелёным (если > 0)
        likes_col = headers.index("Likes") + 1
        ws.conditional_formatting.add(
            f"{ws.cell(row=2, column=likes_col).coordinate}:{ws.cell(row=last_row, column=likes_col).coordinate}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )

        wb.save(filename)


if __name__ == "__main__":
    vk = VKStats(conf.vk_api_key, conf.vk_group_id)

    print(f"Подписчиков сейчас: {vk.get_followers()}")

    START = "2025-09-14"
    END = "2025-09-30"

    rows = vk.get_stats(START, END)
    vk.display_stats(rows)
    vk.save_to_csv(rows, "stats.csv")
    vk.save_to_excel(rows, "stats.xlsx")
    print("\n✅ Статистика сохранена в файлы stats.csv и stats.xlsx")
